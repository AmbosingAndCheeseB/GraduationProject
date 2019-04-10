from sklearn.preprocessing import MinMaxScaler
from openpyxl import load_workbook
from openpyxl import Workbook

load_wb = load_workbook("./Train_Data.xlsx", data_only=True)
load_ws = load_wb['Train_Data']

get_cells = load_ws['G1':'G5919']
value_list = []

for row in get_cells:
    for cell in row:
        value_list.append([cell.value])

scale = MinMaxScaler(feature_range=(0.001, 0.999))
normal = scale.fit_transform(value_list)

write_wb = Workbook()
write_ws = write_wb.active

i = 1
for row in normal:
    for value in row:
        write_ws.cell(i, 1, value)
        i = i+1

write_wb.save('./Normal_Data.xlsx')
